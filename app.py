import streamlit as st
import pandas as pd
import zipfile
import io
import openpyxl.styles
# Ajoutez cette ligne avec les autres imports
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="Fusion et analyse de fichiers CSV",
    page_icon="🔗",
    layout="wide"
)

def upload_files():
    """Composant pour uploader des fichiers ZIP"""
    uploaded_files = st.file_uploader(
        "Glissez-déposez vos fichiers ZIP ici",
        type=['zip'],
        accept_multiple_files=True,
        help="Uploadez un ou plusieurs fichiers ZIP contenant des fichiers CSV"
    )
    return uploaded_files

def merge_csv_files(uploaded_files):
    """Fusionne les fichiers CSV extraits des fichiers ZIP"""
    all_dataframes = []
    
    for uploaded_file in uploaded_files:
        try:
            with zipfile.ZipFile(uploaded_file, 'r') as zip_ref:
                csv_files = [f for f in zip_ref.namelist() if f.endswith('.csv')]
                
                for csv_file in csv_files:
                    with zip_ref.open(csv_file) as file:
                        content = file.read()
                        # Lire avec l'encodage ISO-8859-1 et séparateur ;
                        df = pd.read_csv(
                            io.BytesIO(content),
                            encoding='ISO-8859-1',
                            sep=';',
                            quotechar='"',
                            dtype=str  # Garder tous les types comme string
                        )
                        all_dataframes.append(df)
                        st.info(f"✅ Fichier traité : {csv_file} ({len(df)} lignes)")
                        
        except Exception as e:
            st.error(f"❌ Erreur lors du traitement de {uploaded_file.name}: {str(e)}")
    
    if all_dataframes:
        merged_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
        return merged_df
    else:
        return None

def analyser_consommation_data(df):
    """Analyse la consommation de données et prépare la feuille 'Moyenne conso DATA'"""
    # Filtrer les lignes dont "Nom de la sous-rubrique" commence par "Echanges"
    if "Nom de la sous-rubrique" not in df.columns:
        st.error("❌ Colonne 'Nom de la sous-rubrique' manquante dans les données")
        return None
    
    # Corriger les valeurs NaN avant d'appliquer startswith
    df = df.copy()  # Créer une copie explicite du DataFrame
    df["Nom de la sous-rubrique"] = df["Nom de la sous-rubrique"].fillna("")
    filtered_df = df[df["Nom de la sous-rubrique"].str.startswith("Echanges")].copy()  # Créer une copie explicite
    
    if filtered_df.empty:
        st.warning("⚠️ Aucune donnée commençant par 'Echanges' trouvée")
        return None
    
    # Extraire les périodes de facturation et convertir en format date
    if "Période de la facture" not in filtered_df.columns:
        st.error("❌ Colonne 'Période de la facture' manquante dans les données")
        return None
    
    # Convertir la colonne de période au format date en utilisant .loc
    filtered_df.loc[:, "Date de facturation"] = pd.to_datetime(
        filtered_df["Période de la facture"],
        format="%d/%m/%Y", 
        errors="coerce"
    )
    
    # Créer une colonne pour le tri (format AAAAMM)
    filtered_df.loc[:, "Date_tri"] = filtered_df["Date de facturation"].dt.strftime("%Y%m")
    
    # Extraire le mois et l'année au format "mois-aa"
    filtered_df.loc[:, "Mois-Année"] = filtered_df["Date de facturation"].dt.strftime("%B-%y")
    
    # Traduire les noms des mois en français
    mois_fr = {
        'January': 'janvier', 'February': 'février', 'March': 'mars', 
        'April': 'avril', 'May': 'mai', 'June': 'juin',
        'July': 'juillet', 'August': 'août', 'September': 'septembre', 
        'October': 'octobre', 'November': 'novembre', 'December': 'décembre'
    }
    
    # Appliquer la traduction
    for en, fr in mois_fr.items():
        filtered_df.loc[:, "Mois-Année"] = filtered_df["Mois-Année"].str.replace(en, fr)
    
    # Colonnes fixes requises
    required_cols = [
        "Nom de la rubrique de niveau 1",
        "Numéro de l'utilisateur",
        "Nom de l'utilisateur",
        "Prénom de l'utilisateur",
        "Numéro de téléphone"
    ]
    
    # Vérifier que toutes les colonnes requises existent
    missing_cols = [col for col in required_cols if col not in filtered_df.columns]
    if missing_cols:
        st.error(f"❌ Colonnes manquantes: {', '.join(missing_cols)}")
        return None
    
    # Colonne pour les volumes
    volume_col = None
    possible_volume_cols = ["Volume consommé", "Volume Data", "Volume", "Quantité ou volume"]
    for col in possible_volume_cols:
        if col in filtered_df.columns:
            volume_col = col
            break
    
    if not volume_col:
        st.error("❌ Aucune colonne de volume trouvée")
        return None
    
    # Fonctions pour le traitement des volumes
    def parse_volume(vol_str):
        """Convertit une chaîne de volume en Go"""
        try:
            if isinstance(vol_str, (int, float)):
                return vol_str
            
            if pd.isna(vol_str) or vol_str == "":
                return 0
            
            total_bytes = 0
            
            # Gérer différents formats possibles
            if "Go" in vol_str and "Mo" in vol_str and "Ko" in vol_str:
                parts = vol_str.split()
                go_idx = parts.index("Go")
                mo_idx = parts.index("Mo")
                ko_idx = parts.index("Ko")
                
                go = float(parts[go_idx-1]) if go_idx > 0 else 0
                mo = float(parts[mo_idx-1]) if mo_idx > 0 else 0
                ko = float(parts[ko_idx-1]) if ko_idx > 0 else 0
                
                total_bytes = go * 1024 * 1024 * 1024 + mo * 1024 * 1024 + ko * 1024
            
            elif "Go" in vol_str and "Mo" in vol_str:
                parts = vol_str.split()
                go_idx = parts.index("Go")
                mo_idx = parts.index("Mo")
                
                go = float(parts[go_idx-1]) if go_idx > 0 else 0
                mo = float(parts[mo_idx-1]) if mo_idx > 0 else 0
                
                total_bytes = go * 1024 * 1024 * 1024 + mo * 1024 * 1024
            
            elif "Go" in vol_str:
                parts = vol_str.split()
                go_idx = parts.index("Go")
                go = float(parts[go_idx-1]) if go_idx > 0 else 0
                total_bytes = go * 1024 * 1024 * 1024
            
            elif "Mo" in vol_str:
                parts = vol_str.split()
                mo_idx = parts.index("Mo")
                mo = float(parts[mo_idx-1]) if mo_idx > 0 else 0
                total_bytes = mo * 1024 * 1024
                
            elif "Ko" in vol_str:
                parts = vol_str.split()
                ko_idx = parts.index("Ko")
                ko = float(parts[ko_idx-1]) if ko_idx > 0 else 0
                total_bytes = ko * 1024
            
            else:
                try:
                    total_bytes = float(vol_str)
                except:
                    return 0
                    
            return total_bytes / (1024 * 1024 * 1024)  # Conversion en Go
        except:
            return 0
    
    def format_volume(vol_go):
        """Convertit un volume en Go en format texte avec 2 décimales"""
        try:
            if pd.isna(vol_go) or vol_go == 0:
                return "0.00 Go"
                
            return f"{vol_go:.2f} Go"
        except:
            return "0.00 Go"
    
    # Appliquer la conversion en utilisant .loc
    filtered_df.loc[:, "Volume_Go"] = filtered_df[volume_col].apply(parse_volume)
    
    # Création d'une table de correspondance entre les mois-années et leurs dates de tri
    month_order = filtered_df.drop_duplicates("Mois-Année")[["Mois-Année", "Date_tri"]]
    month_order = dict(zip(month_order["Mois-Année"], month_order["Date_tri"]))
    
    # Ajouter les colonnes dynamiques par mois (pivot)
    pivot_df = filtered_df.pivot_table(
        index=required_cols,
        columns="Mois-Année",
        values="Volume_Go",
        aggfunc="sum",
        fill_value=0
    ).reset_index()
    
    # Trier les colonnes de mois du plus récent au plus ancien
    month_cols = [col for col in pivot_df.columns if col not in required_cols]
    month_cols = sorted(month_cols, key=lambda x: month_order.get(x, ""), reverse=True)
    pivot_df = pivot_df[required_cols + month_cols]
    
    # Créer la base du dataframe résultat
    result = pivot_df.copy()
    
    # Calculer le total en Go et les moyennes
    result["Volume_Total"] = result[month_cols].sum(axis=1)
    
    # Calculer les moyennes
    result["Volume_Moy_Total"] = result[month_cols].mean(axis=1)
    
    if len(month_cols) >= 4:
        result["Volume_Moy_4_Mois"] = result[month_cols[:4]].mean(axis=1)
    else:
        result["Volume_Moy_4_Mois"] = result["Volume_Moy_Total"]
    
    # AJOUT : Trier le dataframe par volume total décroissant
    result = result.sort_values(by="Volume_Total", ascending=False)
    
    # Modifications apportées ici
    # Ne pas formater les colonnes, garder les valeurs numériques
    result["Total (Go)"] = result["Volume_Total"]
    result["Moyenne (Go) 4 mois"] = result["Volume_Moy_4_Mois"]
    result["Moyenne (Go) total"] = result["Volume_Moy_Total"]
    
    # Réorganiser les colonnes dans l'ordre final demandé
    final_cols = required_cols + ["Total (Go)", "Moyenne (Go) 4 mois", "Moyenne (Go) total"] + month_cols
    result = result[final_cols]
    
    # Supprimer les colonnes temporaires
    result = result.drop(columns=["Volume_Total", "Volume_Moy_Total", "Volume_Moy_4_Mois"], errors="ignore")
    
    return result

# Remplacer la fonction parse_volume_text par celle-ci
def parse_volume_text(vol_str):
    """Convertit une chaîne de volume formatée en Go (valeur numérique)"""
    try:
        # Cas des valeurs déjà numériques
        if isinstance(vol_str, (int, float)):
            return float(vol_str)
        
        if pd.isna(vol_str) or vol_str == "":
            return 0
        
        # Pour une meilleure approche, utiliser la même logique que dans parse_volume
        if "Go" in vol_str or "Mo" in vol_str or "Ko" in vol_str:
            go, mo, ko = 0, 0, 0
            
            parts = vol_str.split()
            for i in range(0, len(parts) - 1, 2):
                try:
                    value = float(parts[i])
                    unit = parts[i + 1]
                    
                    if unit == "Go":
                        go = value
                    elif unit == "Mo":
                        mo = value
                    elif unit == "Ko":
                        ko = value
                except:
                    pass
            
            # Convertir tout en Go
            total_go = go + mo / 1024 + ko / (1024 * 1024)
            return total_go
        else:
            try:
                return float(vol_str)
            except:
                return 0
    
    except Exception as e:
        return 0

# Remplacer la fonction create_excel_file par celle-ci
def create_excel_file(dataframe, analysis_df=None):
    """
    Crée un fichier Excel contenant toutes les feuilles d'analyse
    """
    excel_buffer = io.BytesIO()
    
    with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
        # 1. Feuille d'export standard (toujours en premier)
        dataframe.to_excel(
            writer,
            sheet_name='Export',
            index=False
        )
        
        # 2. Feuille d'analyse de consommation DATA (si disponible)
        if analysis_df is not None and not analysis_df.empty:
            # Créer une copie du DataFrame pour la manipulation
            export_df = analysis_df.copy()
            
            # Identifier toutes les colonnes de volume (fixes et mensuelles)
            fixed_cols = ["Nom de la rubrique de niveau 1", "Numéro de l'utilisateur", 
                         "Nom de l'utilisateur", "Prénom de l'utilisateur", "Numéro de téléphone"]
            all_volume_cols = [col for col in export_df.columns if col not in fixed_cols]
            
            # Créer un nouveau DataFrame pour l'export avec valeurs numériques
            numeric_df = export_df.copy()
            
            # Convertir les colonnes de volume en valeurs numériques
            for col in all_volume_cols:
                numeric_df[col] = numeric_df[col].apply(parse_volume_text)
            
            # Calculer les totaux pour chaque colonne numérique
            totals = {}
            for col in all_volume_cols:
                totals[col] = numeric_df[col].sum()
            
            # Créer une ligne de total
            total_row = {col: "" for col in numeric_df.columns}
            for col in fixed_cols:
                if col == "Nom de l'utilisateur":
                    total_row[col] = "TOTAL"
                else:
                    total_row[col] = ""
            
            for col in all_volume_cols:
                total_row[col] = totals[col]
            
            # Ajouter la ligne de total au DataFrame
            numeric_df = pd.concat([numeric_df, pd.DataFrame([total_row])], ignore_index=True)
            
            # Exporter le DataFrame numérique vers Excel
            numeric_df.to_excel(
                writer,
                sheet_name='Moyenne conso DATA',
                index=False
            )
            
            # Récupérer le classeur et la feuille de travail
            workbook = writer.book
            worksheet = writer.sheets['Moyenne conso DATA']
            
            # Obtenir les dimensions de la table
            num_rows = len(numeric_df)
            num_cols = len(numeric_df.columns)
            table_range = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
            
            # Importer les classes nécessaires pour les tableaux Excel
            from openpyxl.worksheet.table import Table, TableStyleInfo
            
            # Créer un tableau et y appliquer un style
            table = Table(displayName="ConsommationData", ref=table_range)
            
            # Appliquer le style "TableStyleMedium16" (bleu moyen 16)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium16",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            
            # Ajouter le tableau à la feuille de travail
            worksheet.add_table(table)
            
            # Figer les volets en cellule F2
            worksheet.freeze_panes = 'F2'
            
            # Appliquer un format simple pour afficher les volumes en Go avec 2 décimales
            go_format = '0.00" Go"'
            
            # Appliquer le format aux colonnes de volume
            for col_idx, col_name in enumerate(numeric_df.columns):
                if col_name not in fixed_cols:
                    # Appliquer le format à toutes les cellules de la colonne y compris la ligne total
                    for row in range(2, len(numeric_df) + 2):
                        cell = worksheet.cell(row=row, column=col_idx + 1)
                        cell.number_format = go_format
            
            # Mettre en évidence la ligne de total avec une police en gras
            for col_idx in range(1, num_cols + 1):
                cell = worksheet.cell(row=num_rows + 1, column=col_idx)
                cell.font = openpyxl.styles.Font(bold=True)
            
            # Ajuster automatiquement la largeur des colonnes
            for col_idx, column in enumerate(numeric_df):
                column_letter = get_column_letter(col_idx + 1)
                # Calculer la largeur maximale en fonction du contenu
                max_length = 0
                # Vérifier l'en-tête
                header_length = len(str(column)) + 2  # +2 pour un peu d'espace supplémentaire
                max_length = max(max_length, header_length)
                
                # Vérifier toutes les valeurs dans la colonne
                for i in range(len(numeric_df)):
                    cell_value = str(numeric_df.iloc[i, col_idx])
                    if col_idx < len(fixed_cols):
                        # Pour les colonnes textuelles
                        cell_length = len(cell_value) + 2
                    else:
                        # Pour les colonnes numériques (format X.XX Go)
                        cell_length = 10  # Taille standard pour les valeurs en Go
                    max_length = max(max_length, cell_length)
                
                # Définir la largeur de la colonne avec une limite maximale
                max_length = min(max_length, 40)  # Limiter à 40 pour éviter des colonnes trop larges
                worksheet.column_dimensions[column_letter].width = max_length
            
            # Adapter la hauteur de ligne pour l'en-tête
            worksheet.row_dimensions[1].height = 25

        # 3. NOUVELLE FEUILLE "Résumé par période" - traitée APRÈS la feuille "Moyenne conso DATA"
        # Ajoutons du logging pour débugger
        print("Création de la feuille Résumé par période")
        
        # Vérifier que dataframe n'est pas vide
        if not dataframe.empty:
            print(f"DataFrame source contient {len(dataframe)} lignes")
            
            # Filtrer les données pour "Depuis le mobile" et "Communications incluses" ou "Communications facturées"
            mask_niveau1 = dataframe["Nom de la rubrique de niveau 1"].str.contains("Depuis le mobile", na=False)
            mask_niveau2 = dataframe["Nom de la rubrique de niveau 2"].str.contains("Communications incluses|Communications facturées", na=False, regex=True)
            resume_df = dataframe[mask_niveau1 & mask_niveau2].copy()
            
            print(f"Après filtrage: {len(resume_df)} lignes")
            
            if not resume_df.empty:
                # Identifier les colonnes de période en cherchant les paires "Qté facturées" / "Montant Facturé"
                period_pairs = []
                headers = list(dataframe.columns)
                
                print(f"En-têtes disponibles: {headers}")
                
                for i in range(len(headers)):
                    if "Qté facturées" in headers[i] and i+1 < len(headers) and "Montant Facturé" in headers[i+1]:
                        # La période est le préfixe commun (comme "mai-25 ")
                        period = headers[i].split("Qté facturées")[0].strip()
                        if period:
                            period_pairs.append((period, headers[i], headers[i+1]))
                
                print(f"Paires de périodes identifiées: {period_pairs}")
                
                if period_pairs:
                    # Créer un DataFrame pour le résumé
                    resume_columns = ["Nom de la rubrique de niveau 1", "Nom de la rubrique de niveau 2", "Nom de la sous-rubrique"]
                    for period, _, _ in period_pairs:
                        resume_columns.append(f"{period} Qté facturées")
                        resume_columns.append(f"{period} Montant Facturé")
                    
                    resume_summary = pd.DataFrame(columns=resume_columns)
                    
                    # Remplir avec les données filtrées
                    for _, row in resume_df.iterrows():
                        new_row = {
                            "Nom de la rubrique de niveau 1": row["Nom de la rubrique de niveau 1"],
                            "Nom de la rubrique de niveau 2": row["Nom de la rubrique de niveau 2"],
                            "Nom de la sous-rubrique": row["Nom de la sous-rubrique"]
                        }
                        
                        for period, qty_col, amount_col in period_pairs:
                            new_row[f"{period} Qté facturées"] = row[qty_col] if qty_col in row else ""
                            new_row[f"{period} Montant Facturé"] = row[amount_col] if amount_col in row else ""
                        
                        resume_summary = pd.concat([resume_summary, pd.DataFrame([new_row])], ignore_index=True)
                    
                    # Exporter vers Excel
                    resume_summary.to_excel(
                        writer,
                        sheet_name='Résumé par période',
                        index=False
                    )
                    
                    # Récupérer la feuille
                    worksheet = writer.sheets['Résumé par période']
                    
                    # Obtenir les dimensions de la table
                    num_rows = len(resume_summary)
                    num_cols = len(resume_summary.columns)
                    table_range = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
                    
                    # Créer un tableau avec style
                    from openpyxl.worksheet.table import Table, TableStyleInfo
                    
                    # Vérifier si le nom de table existe déjà
                    table_name = "ResumePeriode"
                    i = 1
                    while table_name in [t.name for t in workbook._tables.values()]:
                        table_name = f"ResumePeriode{i}"
                        i += 1
                    
                    # Créer un tableau et y appliquer un style
                    table = Table(displayName=table_name, ref=table_range)
                    
                    # Appliquer le style "TableStyleMedium16" (bleu moyen 16)
                    table.tableStyleInfo = TableStyleInfo(
                        name="TableStyleMedium16",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    
                    # Ajouter le tableau à la feuille de travail
                    worksheet.add_table(table)
                    
                    # Figer les volets en cellule D2
                    worksheet.freeze_panes = 'D2'
                    
                    # Mettre en forme les colonnes de montant avec le format monétaire
                    for col_idx in range(1, num_cols + 1):
                        col_name = resume_summary.columns[col_idx - 1]
                        if "Montant Facturé" in col_name:
                            for row_idx in range(2, num_rows + 2):  # +2 car Excel est indexé à 1 et en-têtes
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                cell.number_format = '#,##0.00 €'
                        elif "Qté facturées" in col_name:
                            for row_idx in range(2, num_rows + 2):
                                cell = worksheet.cell(row=row_idx, column=col_idx)
                                # Garder le format nombre standard pour les quantités
                                cell.number_format = '#,##0.00'
                    
                    # Ajuster automatiquement la largeur des colonnes
                    for col_idx, column in enumerate(resume_summary):
                        column_letter = get_column_letter(col_idx + 1)
                        # Largeur basée sur l'en-tête ou contenu
                        max_length = len(str(column)) + 2
                        
                        # Pour les colonnes de base
                        if col_idx < 3:  # Les trois premières colonnes (noms des rubriques)
                            for i in range(len(resume_summary)):
                                cell_value = str(resume_summary.iloc[i, col_idx])
                                max_length = max(max_length, len(cell_value) + 2)
                            # Limiter les colonnes de texte
                            max_length = min(max_length, 40)
                        else:
                            # Pour les colonnes de valeurs (quantités et montants)
                            max_length = 15  # Largeur standard pour les valeurs numériques
                        
                        worksheet.column_dimensions[column_letter].width = max_length
                    
                    # Adapter la hauteur de ligne pour l'en-tête
                    worksheet.row_dimensions[1].height = 25
                    
                    # Ajouter une ligne total à la fin
                    # Nous allons calculer le total pour chaque colonne numérique
                    total_row_idx = num_rows + 2
                    
                    # Écrire "TOTAL" dans la première colonne
                    worksheet.cell(row=total_row_idx, column=1).value = "TOTAL"
                    worksheet.cell(row=total_row_idx, column=1).font = openpyxl.styles.Font(bold=True)
                    
                    # Calculer et écrire les totaux pour les colonnes numériques
                    for col_idx in range(4, num_cols + 1):  # À partir de la 4ème colonne (les valeurs)
                        col_letter = get_column_letter(col_idx)
                        # Formule pour calculer la somme de la colonne
                        formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
                        cell = worksheet.cell(row=total_row_idx, column=col_idx)
                        cell.value = formula
                        cell.font = openpyxl.styles.Font(bold=True)
                        
                        # Appliquer le format approprié (monétaire ou nombre)
                        if "Montant Facturé" in resume_summary.columns[col_idx - 1]:
                            cell.number_format = '#,##0.00 €'
                        else:
                            cell.number_format = '#,##0.00'
    
    excel_buffer.seek(0)
    return excel_buffer

# Ajoutez cette fonction pour créer la feuille Résumé par période
def create_resume_periode_sheet(dataframe, writer):
    """
    Crée la feuille 'Résumé par période' dans le fichier Excel
    """
    import pandas as pd
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo
    import openpyxl.styles
    
    # Filtrer les données pour "Depuis le mobile" et "Communications incluses" ou "Communications facturées"
    mobile_mask = dataframe["Nom de la rubrique de niveau 1"].str.contains("Depuis le mobile", na=False)
    comm_mask = dataframe["Nom de la rubrique de niveau 2"].str.contains("Communications incluses|Communications facturées", na=False, regex=True)
    resume_df = dataframe[mobile_mask & comm_mask].copy()
    
    # Si aucune donnée ne correspond au filtre, sortir sans créer la feuille
    if resume_df.empty:
        return False
    
    # Identifier les colonnes de base et les colonnes de période
    base_cols = ["Nom de la rubrique de niveau 1", "Nom de la rubrique de niveau 2", "Nom de la sous-rubrique"]
    period_cols = []
    
    # Rechercher les colonnes contenant "Qté facturées" et "Montant Facturé"
    qty_cols = [col for col in dataframe.columns if "Qté facturées" in col]
    amount_cols = [col for col in dataframe.columns if "Montant Facturé" in col]
    
    # Créer le nouveau DataFrame avec les colonnes de base
    result_df = resume_df[base_cols].copy()
    
    # Ajouter les colonnes pour les périodes
    for qty_col in qty_cols:
        # Extraire la période (comme "mai-25 ")
        period = qty_col.split("Qté facturées")[0].strip()
        
        # Trouver la colonne de montant correspondante
        amount_col = [col for col in amount_cols if col.startswith(period)][0] if [col for col in amount_cols if col.startswith(period)] else None
        
        if period and amount_col:
            # Ajouter les colonnes au DataFrame résultat
            result_df[f"{period} Qté facturées"] = resume_df[qty_col]
            result_df[f"{period} Montant Facturé"] = resume_df[amount_col]
            
            # Garder trace des paires de colonnes
            period_cols.append((period, f"{period} Qté facturées", f"{period} Montant Facturé"))
    
    # Exporter vers Excel
    result_df.to_excel(
        writer,
        sheet_name='Résumé par période',
        index=False
    )
    
    # Récupérer la feuille de travail
    worksheet = writer.sheets['Résumé par période']
    
    # Obtenir les dimensions de la table
    num_rows = len(result_df)
    num_cols = len(result_df.columns)
    
    # Appliquer le style de tableau
    table_range = f"A1:{get_column_letter(num_cols)}{num_rows + 1}"
    
    try:
        # Créer un tableau et y appliquer un style
        table = Table(displayName="ResumePeriode", ref=table_range)
        
        # Appliquer le style "TableStyleMedium16" (bleu moyen 16)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium16",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        
        # Ajouter le tableau à la feuille de travail
        worksheet.add_table(table)
    except Exception as e:
        # En cas d'erreur, continuer sans ajouter le tableau
        print(f"Erreur lors de la création du tableau: {e}")
    
    # Figer les volets en cellule D2
    worksheet.freeze_panes = 'D2'
    
    # Mettre en forme les colonnes de montant avec le format monétaire
    for col_idx in range(1, num_cols + 1):
        col_name = result_df.columns[col_idx - 1]
        if "Montant Facturé" in col_name:
            for row_idx in range(2, num_rows + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '#,##0.00 €'
        elif "Qté facturées" in col_name:
            for row_idx in range(2, num_rows + 2):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.number_format = '#,##0.00'
    
    # Ajuster automatiquement la largeur des colonnes
    for col_idx, column in enumerate(result_df):
        column_letter = get_column_letter(col_idx + 1)
        
        # Largeur basée sur l'en-tête ou contenu
        if col_idx < len(base_cols):  # Colonnes de texte
            max_length = len(str(column)) + 2
            for i in range(len(result_df)):
                cell_value = str(result_df.iloc[i, col_idx])
                max_length = max(max_length, len(cell_value) + 2)
            max_length = min(max_length, 40)
        else:  # Colonnes de valeurs
            max_length = 15
        
        worksheet.column_dimensions[column_letter].width = max_length
    
    # Adapter la hauteur de ligne pour l'en-tête
    worksheet.row_dimensions[1].height = 25
    
    # Ajouter une ligne total
    total_row_idx = num_rows + 2
    
    # Écrire "TOTAL" dans la première colonne
    worksheet.cell(row=total_row_idx, column=1).value = "TOTAL"
    worksheet.cell(row=total_row_idx, column=1).font = openpyxl.styles.Font(bold=True)
    
    # Calculer et écrire les totaux pour les colonnes numériques (à partir de la 4ème colonne)
    for col_idx in range(4, num_cols + 1):
        col_letter = get_column_letter(col_idx)
        formula = f"=SUM({col_letter}2:{col_letter}{total_row_idx-1})"
        cell = worksheet.cell(row=total_row_idx, column=col_idx)
        cell.value = formula
        cell.font = openpyxl.styles.Font(bold=True)
        
        # Appliquer le format approprié
        if "Montant Facturé" in result_df.columns[col_idx - 1]:
            cell.number_format = '#,##0.00 €'
        else:
            cell.number_format = '#,##0.00'
    
    return True


def main():
    st.title("🔗 Fusion et analyse de fichiers CSV")
    st.markdown("---")
    
    # Zone de dépôt des fichiers
    uploaded_files = upload_files()
    
    if uploaded_files:
        st.success(f"{len(uploaded_files)} fichier(s) uploadé(s)")
        
        # Bouton pour démarrer la fusion
        if st.button("🚀 Démarrer la fusion et l'analyse", type="primary"):
            with st.spinner("Traitement en cours..."):
                try:
                    # Fusion des fichiers
                    merged_df = merge_csv_files(uploaded_files)
                    
                    if merged_df is not None and not merged_df.empty:
                        st.success(f"✅ Fusion terminée ! {len(merged_df)} lignes fusionnées")
                        
                        # Analyse des données pour la feuille "Moyenne conso DATA"
                        with st.spinner("Analyse des consommations DATA en cours..."):
                            analysis_df = analyser_consommation_data(merged_df)
                            
                            if analysis_df is not None and not analysis_df.empty:
                                st.success(f"✅ Analyse terminée ! {len(analysis_df)} utilisateurs analysés")
                            else:
                                st.warning("⚠️ Impossible de créer l'analyse des consommations DATA. Vérifiez le format des données.")
                        
                        # Création du fichier Excel avec toutes les feuilles
                        excel_buffer = create_excel_file(merged_df, analysis_df)
                        
                        # Un seul bouton de téléchargement
                        st.download_button(
                            label="📥 Télécharger Analyse de Parc.xlsx",
                            data=excel_buffer,
                            file_name="Analyse de Parc.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    else:
                        st.error("❌ Aucune donnée à fusionner")
                        
                except Exception as e:
                    st.error(f"❌ Erreur lors du traitement : {str(e)}")
                    st.exception(e)

if __name__ == "__main__":
    main()